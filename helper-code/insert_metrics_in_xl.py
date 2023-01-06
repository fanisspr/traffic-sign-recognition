import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from typing import Iterable, Iterator, Match, Tuple


def get_metrics(matches: Iterator[Match[str]]) -> list[float]:
    '''
    Gets metric data from matches and splits it to training and testing
    '''
    training = []
    testing = []
    for train, test in zip(matches, matches):
        training.append(float(train.group()))
        testing.append(float(test.group()))
    return training, testing


def put_in_col(data: Iterable, starting_cell: str) -> None:
    '''
    Writes the metric data in a column starting from starting cell 
    '''
    char = starting_cell[0]
    for i, row in enumerate(range(int(starting_cell[1:]), int(starting_cell[1:]) + len(data))):
        value = data[i]
        cell = ws[char + str(row)]
        if type(value) is int:
            cell.number_format = numbers.FORMAT_GENERAL
        else:
            cell.number_format = '0.000;' if value > 1e-3 else '0.000E+00;'
        cell.value = data[i]


def set_headers(optimizer: str, offset: Tuple[int, int]) -> None:
    '''
    Write the headers of each column
    '''
    # 1st column:
    char = get_column_letter(1 + offset[1])
    # optimizer
    ws[char + str(1 + offset[0])] = f'{opt}'
    # epochs
    ws[char + str(2 + offset[0])] = 'epochs'

    # Training metrics:
    # 2nd column
    char = get_column_letter(2 + offset[1])
    ws[char + str(1 + offset[0])] = 'Training'
    ws[char + str(2 + offset[0])] = f'{optimizer} loss'
    # 3d column
    char = get_column_letter(3 + offset[1])
    ws[char + str(2 + offset[0])] = f'{optimizer} acc'

    # Testing metrics:
    # 4th column
    char = get_column_letter(4 + offset[1])
    ws[char + str(1 + offset[0])] = 'Testing'
    ws[char + str(2 + offset[0])] = f'{optimizer} loss'
    # 5th column
    char = get_column_letter(5 + offset[1])
    ws[char + str(2 + offset[0])] = f'{optimizer} acc'


def fill_sheet(title: str):
    pass


os.chdir(r'F:\Φάνης\Desktop\διπλωματικη')
print(os.getcwd())

# def main():
filename = "GTSRB.xlsx"
data_dir = 'tsr5.3-Vit_SPT_LSA'
logs = 'training.txt'
normalization = None  # Choose between: BatchNorm, LRN, None
# sheetname = f'{data_dir}_{normalization}'
sheetname = f'{data_dir}'
wb = load_workbook(filename)
# book.get_sheet_by_name('someWorksheetName')
if sheetname not in wb.sheetnames:
    wb.create_sheet(sheetname)
ws = wb[sheetname]

optimizers = ['Adam', 'Rmsprop', 'SGD', 'SGDW']

offset = [(0, 0), (0, 6), (28, 0), (28, 6)]
for opt, off in zip(optimizers, offset):
    if normalization:
        with open(f"{data_dir}/{opt}/{normalization}/{logs}", "r") as f:
            txt = f.read()
    else:
        with open(f"{data_dir}/{opt}/{logs}", "r") as f:
            txt = f.read()

    pattern = re.compile(r'(?<=loss: ).{5,6}(e-\d+)?')
    matches = pattern.finditer(txt)

    # other optimizers have training/testing reversed after tensorflow update
    if opt == 'Adam' and data_dir == 'tsr4.3-Vit':
        training_loss, testing_loss = get_metrics(matches)
    else:
        testing_loss, training_loss = get_metrics(matches)

    pattern = re.compile(r'(?<=accuracy: ).{5,6}(e-\d+)?')
    matches = pattern.finditer(txt)

    if opt == 'Adam' and data_dir == 'tsr4.3-Vit':
        training_acc, testing_acc = get_metrics(matches)
    else:
        testing_acc, training_acc = get_metrics(matches)

    set_headers(opt, off)

    # epochs
    char = get_column_letter(1 + off[1])
    put_in_col(range(1, 26), char + str(3 + off[0]))

    # B3:E27 - metrics
    metrics = [training_loss, training_acc, testing_loss, testing_acc]
    for metric, col in zip(metrics, range(2 + off[1], 2 + off[1] + 4)):
        char = get_column_letter(col)
        put_in_col(metric, char + str(3 + off[0]))

wb.save(filename)

print(
    f'Inserted metrics in {filename} from folder {data_dir}')
# if __name__ == '__main__':
#     main()
